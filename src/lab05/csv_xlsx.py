import sys, os, csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "lib")))
from pathProccesing import pathProccesing


def csv_to_xlsx(csv_path: str, xlsx_path: str, path_suffix=(".csv", ".xlsx")) -> None:
    csv_path, xlsx_path = pathProccesing(csv_path, False, path_suffix), pathProccesing(
        xlsx_path, True, path_suffix
    )
    if csv_path:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        with csv_path.open(encoding="utf-8") as f_csv:
            read_csv = csv.DictReader(f_csv)
            rows = [{headers: headers for headers in read_csv.fieldnames}] + list(
                read_csv
            )
            max_len_w = max(
                [len(val) for r_val in rows for val in list(r_val.values())] + [8]
            )
            for r in rows:
                ws.append(list(r.values()))
            for index in range(1, len(read_csv.fieldnames) + 1):
                ws.column_dimensions[get_column_letter(index)].width = max_len_w
            wb.save(xlsx_path)
    else:
        return csv_path


csv_to_xlsx("data/lab05/samples/cities.csv", "data/lab05/out/people.xlsx")
