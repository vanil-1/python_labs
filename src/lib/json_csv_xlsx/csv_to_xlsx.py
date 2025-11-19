import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

def path_proccesing(path_str: str, out: bool):
    path = Path(path_str)
    if out: path.parent.mkdir(parents = True, exist_ok = True)
    return path
def path_is_csv(path: Path, error = None):
    if path.exists():
        if path.suffix.lower() == ".csv": error = False
        else: error = ValueError
    else: error = FileNotFoundError
    return error

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_path, xlsx_path = path_proccesing(csv_path, False), path_proccesing(xlsx_path, True) 
    if not(path_is_csv(csv_path)):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        with csv_path.open(encoding = "utf-8") as f_csv:
            read_csv = csv.DictReader(f_csv)
            rows = [{headers: headers for headers in read_csv.fieldnames}] + list(read_csv)
            max_len_w = max([len(val) for r_val in rows for val in list(r_val.values())] + [8])
            for r in rows: ws.append(list(r.values()))
            for index in range(1, len(read_csv.fieldnames) + 1): ws.column_dimensions[get_column_letter(index)].width = max_len_w
            wb.save(xlsx_path)
    else: return path_is_csv(csv_path)