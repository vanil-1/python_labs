import csv, json, pathlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def pathProccesing(path: str, path_out: bool, path_suffix: tuple|list|str):
    path = pathlib.Path(path)
    if path.suffix.lower() in path_suffix:
        if path.exists(): return path
        elif not(path.exists()) and path_out:
            path.parent.mkdir(parents = True, exist_ok = True)
            return path
        else: return FileNotFoundError
    else: return ValueError

def csv_to_json(csv_path: str, json_path: str) -> None:
    csv_path, json_path = pathProccesing(csv_path, False, '.csv'), pathProccesing(json_path, True, '.json')
    try:
        with json_path.open("w", newline = "", encoding = "utf-8") as f_json, csv_path.open(encoding = "utf-8") as f_csv:
            read_csv = csv.DictReader(f_csv)
            rows = [{headers: headers for headers in read_csv.fieldnames}] + list(read_csv)
            json.dump(rows, f_json, ensure_ascii = False, indent = 2)
        print('Succses...')
    except: print(csv_path, json_path)

def json_to_csv(json_path: str, csv_path: str) -> None:
    json_path, csv_path = pathProccesing(json_path, False, '.json'), pathProccesing(csv_path, True, '.csv') 
    try:
        with csv_path.open("w", newline = "", encoding = "utf-8") as f_csv, json_path.open(encoding = "utf-8") as f_json:
            rows = json.load(f_json)
            w = csv.DictWriter(f_csv, fieldnames = list(rows[0].keys()))
            w.writeheader()
            for r in rows: w.writerow(r)
        print('Succses...')
    except: print(json_path, csv_path)

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    csv_path, xlsx_path = pathProccesing(csv_path, False, '.csv'), pathProccesing(xlsx_path, True, '.xlsx') 
    try:
        if xlsx_path != ValueError:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            with csv_path.open(encoding = "utf-8") as f_csv:
                read_csv = csv.DictReader(f_csv)
                rows = [{headers: headers for headers in read_csv.fieldnames}] + list(read_csv)
                max_len_w = max([len(val) for r_val in rows for val in list(r_val.values())] + [8])
                for r in rows: ws.append(list(r.values()))
                for index in range(1, len(read_csv.fieldnames) + 1): ws.column_dimensions[get_column_letter(index)].width = max_len_w
                wb.save(xlsx_path)
            print('Succses...')
        else: print(xlsx_path)
    except: print(csv_path, xlsx_path)