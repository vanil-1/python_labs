import csv, json, pathlib
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def pathProccesing(path: str, path_out: bool, path_suffix: tuple | list | str):
    path = pathlib.Path(path)
    if path.suffix.lower() in path_suffix:
        if path.exists():
            return path
        elif not (path.exists()) and path_out:
            path.parent.mkdir(parents=True, exist_ok=True)
            return path
        else:
            return FileNotFoundError
    else:
        return ValueError


def json_csv_xlsx(
    path_in: str, path_out: str, path_suffix=(".json", ".csv", ".xlsx")
) -> None:
    path_in, path_out = pathProccesing(path_in, False, path_suffix), pathProccesing(
        path_out, True, path_suffix
    )
    path_in_suffix, path_out_suffix = path_in.suffix.lower(), path_out.suffix.lower()
    try:
        with path_out.open("w", newline="", encoding="utf-8") as f_out, path_in.open(
            encoding="utf-8"
        ) as f_in:

            if path_in_suffix == ".csv" and path_out_suffix == ".json":
                read_csv = csv.DictReader(f_in)
                rows = [{headers: headers for headers in read_csv.fieldnames}] + list(
                    read_csv
                )
                json.dump(rows, f_out, ensure_ascii=False, indent=2)

            elif path_in_suffix == ".json" and path_out_suffix == ".csv":
                rows = json.load(f_in)
                w = csv.DictWriter(f_out, fieldnames=list(rows[0].keys()))
                w.writeheader()
                for r in rows:
                    w.writerow(r)

            elif path_in_suffix == ".csv" and path_out_suffix == ".xlsx":
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"

                read_csv = csv.DictReader(f_in)
                rows = [{headers: headers for headers in read_csv.fieldnames}] + list(
                    read_csv
                )
                max_len_w = max(
                    [len(val) for r_val in rows for val in list(r_val.values())] + [8]
                )

                for r in rows:
                    ws.append(list(r.values()))
                for index in range(1, len(read_csv.fieldnames) + 1):
                    ws.column_dimensions[get_column_letter(index)].width = max_len_w

                wb.save(path_out)

        print("Succses...")

    except:
        print(path_in, path_out)
